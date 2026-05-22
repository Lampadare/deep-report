#!/usr/bin/env python3
"""Phase 1: Setup - Directory creation, config, seed processing."""

import os
import re
import sys
import json
import subprocess
import threading
from pathlib import Path
from datetime import datetime
from typing import Optional

from ..state import State
from ..utils import spawn_agent, generate_mcp_config, AGENT_TOOLS
from ..ui import ui

# Extensions we accept as seed references
_SEED_EXTENSIONS = {'.pdf', '.md', '.txt', '.html', '.json', '.xlsx', '.xls', '.csv', '.docx'}


def _is_visible_seed(path: Path) -> bool:
    """Return True if path is a non-hidden file with a supported extension."""
    return (
        path.is_file()
        and not path.name.startswith('.')
        and path.suffix.lower() in _SEED_EXTENSIONS
    )


def _sanitize_topic(topic: str) -> str:
    """Sanitize topic to prevent path traversal and shell injection."""
    import urllib.parse

    # Decode URL-encoded characters first (handles %2e, %2f, etc.)
    try:
        topic = urllib.parse.unquote(topic)
    except Exception:
        pass

    # Remove null bytes
    topic = topic.replace('\x00', '')

    # Normalize unicode to catch homograph attacks
    import unicodedata
    topic = unicodedata.normalize('NFKC', topic)

    # Remove path traversal attempts
    topic = topic.replace("..", "").replace("/", " ").replace("\\", " ")

    # Remove shell metacharacters
    topic = re.sub(r'[;&|`$<>\'\"()]', '', topic)

    # Collapse whitespace
    topic = re.sub(r'\s+', ' ', topic)

    return topic.strip()


def _is_writable(path: Path) -> bool:
    """Check if a directory is writable."""
    try:
        if not path.exists():
            path.mkdir(parents=True, exist_ok=True)
        test_file = path / ".write_test"
        test_file.touch()
        test_file.unlink()
        return True
    except (PermissionError, OSError):
        return False


def _determine_output_dir(args: dict) -> Path:
    """Determine output directory with smart defaults.

    Priority:
    1. Explicit --output flag
    2. cwd if writable
    3. Fallback to ~/Documents/deep-reports/
    """
    topic = args.get("topic", "research")
    slug = _slugify(topic)[:50]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    report_name = f"{slug}_{timestamp}"

    # 1. Explicit --output flag takes priority
    if args.get("report_dir"):
        return Path(args["report_dir"])

    # 2. Default to cwd if writable
    cwd = Path(args.get("cwd") or os.getcwd())
    cwd_output = cwd / report_name
    if _is_writable(cwd):
        return cwd_output

    # 3. Fallback to ~/Documents/deep-reports/
    fallback = Path.home() / "Documents" / "deep-reports" / report_name
    fallback.parent.mkdir(parents=True, exist_ok=True)
    return fallback


def _auto_detect_seeds(cwd: Path) -> Optional[Path]:
    """Auto-detect seed refs folder in cwd.

    Looks for common seed folder names and returns the first non-empty one.
    """
    candidates = ["seed-refs", "seeds", "references", "refs"]
    for name in candidates:
        path = cwd / name
        if path.is_dir():
            files = [f for f in path.iterdir() if _is_visible_seed(f)]
            if files:
                return path
    return None


def run_setup(state: State, args: dict) -> bool:
    """Run the setup phase.

    Args:
        state: The orchestrator state
        args: Dict with keys:
            - topic: Research topic (required)
            - report_dir: Where to create report (optional, defaults to ~/reports/<topic-slug>)
            - model: Research model (sonnet/opus, default sonnet)
            - agent_count: Number of research agents (default 10)
            - seed_refs: Path to folder or list of URLs
            - download_papers: Whether to download cited papers
            - generate_audio: Whether to generate audio version
            - expertise_level: beginner/intermediate/expert
            - report_type: deep-dive/tutorial/comparison/survey

    Returns:
        True if setup succeeded, False otherwise
    """
    # Extract args
    topic = args.get("topic", "").strip()
    if not topic:
        ui.error("Setup failed: no topic provided")
        return False

    # Sanitize topic before any use
    topic = _sanitize_topic(topic)
    if not topic:
        ui.error("Setup failed: topic is empty after sanitization")
        return False

    state.topic = topic
    state.brief = args.get("brief", "")

    # Determine report directory using new logic
    report_dir = _determine_output_dir(args)

    # Validate resolved path is absolute and within allowed parents
    resolved = report_dir.resolve()
    allowed_parents = [Path.home(), Path(args.get("cwd") or os.getcwd()).resolve()]

    def _is_path_under(child: Path, parent: Path) -> bool:
        """Check if child path is under parent using relative_to (not prefix matching)."""
        try:
            child.resolve().relative_to(parent.resolve())
            return True
        except ValueError:
            return False

    if not any(_is_path_under(resolved, p) for p in allowed_parents):
        ui.error("Setup failed: report directory must be within home or current working directory")
        return False

    state.report_dir = str(report_dir)

    # Skill/agent contract: emit REPORT_DIR=<path> on stdout immediately so the
    # driver can find the progress file to tail. This is gated on machine_mode
    # via ui._machine_mode; printed BEFORE any other UI output for this phase
    # so an anchored grep `^REPORT_DIR=` matches reliably.
    if getattr(ui, "_machine_mode", False):
        print(f"REPORT_DIR={report_dir}", flush=True)

    # Set state file path early, before any save() or checkpoint() calls
    state._state_file = str(Path(report_dir) / "state" / "orchestrator_state.json")

    state.current_phase = 1
    state.checkpoint("setup_started")

    # Store config
    state.research_model = args.get("model", "sonnet")
    state.agent_count = min(args.get("agent_count", 10), 30)
    state.download_papers = args.get("download_papers", True)
    state.generate_audio = args.get("generate_audio", False)
    state.expertise_level = args.get("expertise_level", "intermediate")
    state.report_type = args.get("report_type", "deep-dive")
    state.seed_urls = args.get("seed_urls", [])
    state.seed_refs_folder = args.get("seed_refs_folder")
    state.interactive = args.get("interactive", False)

    # Auto-detect seeds if not explicitly provided
    if not state.seed_refs_folder and not state.seed_urls:
        cwd = Path(args.get("cwd") or os.getcwd())
        auto_seeds = _auto_detect_seeds(cwd)
        if auto_seeds:
            ui.info(f"Auto-detected seed refs: {auto_seeds}")
            state.seed_refs_folder = str(auto_seeds)

    state.save()
    state.checkpoint("config_saved")

    # Create directory structure
    ui.step(f"Creating report directory: {report_dir}")
    try:
        _create_directories(report_dir)
    except (OSError, PermissionError) as e:
        ui.error(f"Directory creation failed: {e}")
        return False
    state.checkpoint("directories_created")

    # Register in central registry
    try:
        from ..registry import registry
        registry.register(report_dir, topic)
    except Exception:
        pass  # Registry is non-critical

    # Write initial manifest
    try:
        _write_manifest(state)
    except (OSError, PermissionError) as e:
        ui.error(f"Manifest writing failed: {e}")
        return False
    state.checkpoint("manifest_written")

    # Process seeds if provided
    seeds_to_process = []

    if state.seed_refs_folder:
        folder = Path(state.seed_refs_folder)
        if folder.exists():
            seeds_to_process.extend([str(f) for f in folder.iterdir() if _is_visible_seed(f)])

    if state.seed_urls:
        seeds_to_process.extend(state.seed_urls)

    if seeds_to_process:
        ui.step(f"Processing {len(seeds_to_process)} seed references")
        success = _process_seeds(state, seeds_to_process)
        if not success:
            ui.warning("Seed processing had errors, continuing anyway")
        state.seeds_processed = True
        state.checkpoint("seeds_processed")

        # Summarize seeds
        ui.step("Summarizing seeds")
        _summarize_seeds(state)
        state.seeds_summarized = True
        state.checkpoint("seeds_summarized")

    # Write scope document
    ui.step("Writing scope document")
    _write_scope(state)
    state.scope_written = True
    state.checkpoint("scope_written")

    state.mark_phase_complete(1)
    return True


def _slugify(text: str) -> str:
    """Convert text to URL-safe slug."""
    import re
    text = text.lower()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    slug = text.strip('-')
    # Final safety: use only the basename to prevent any path components
    return os.path.basename(slug) if slug else "report"


def _create_directories(report_dir: Path):
    """Create the full directory structure."""
    dirs = [
        report_dir / "full" / "agents",
        report_dir / "full" / "seeds",
        report_dir / "summaries" / "agents",
        report_dir / "summaries" / "seeds",
        report_dir / "state",
        report_dir / "papers",
    ]
    for d in dirs:
        d.mkdir(parents=True, exist_ok=True)


def _write_manifest(state: State):
    """Write the manifest.json file.

    Raises:
        OSError: If file cannot be written
    """
    manifest = {
        "topic": state.topic,
        "created_at": state.created_at,
        "research_model": state.research_model,
        "agent_count": state.agent_count,
        "download_papers": state.download_papers,
        "generate_audio": state.generate_audio,
        "expertise_level": state.expertise_level,
        "report_type": state.report_type,
        "current_phase": state.current_phase,
        "current_step": state.current_step,
    }
    manifest_path = Path(state.report_dir) / "state" / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2))


def _read_excel_file(path: Path) -> str:
    """Read Excel file and convert to markdown text."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)

        output = []
        for sheet_name in wb.sheetnames:
            output.append(f"\n## Sheet: {sheet_name}\n")
            ws = wb[sheet_name]

            rows = []
            for row in ws.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([str(cell) if cell is not None else '' for cell in row])

            if rows:
                # Create markdown table
                header = rows[0]
                output.append("| " + " | ".join(header) + " |")
                output.append("| " + " | ".join(["---"] * len(header)) + " |")
                for row in rows[1:]:
                    # Pad row if needed
                    while len(row) < len(header):
                        row.append('')
                    output.append("| " + " | ".join(row[:len(header)]) + " |")

        return "\n".join(output)

    except ImportError:
        try:
            import pandas as pd
            df_dict = pd.read_excel(path, sheet_name=None)

            output = []
            for sheet_name, df in df_dict.items():
                output.append(f"\n## Sheet: {sheet_name}\n")
                output.append(df.to_markdown(index=False))

            return "\n".join(output)

        except ImportError:
            return None


def _process_seeds(state: State, seeds: list[str]) -> bool:
    """Process seed references using Claude agents."""
    return _process_seeds_via_agent(state, seeds)


def _process_seeds_via_agent(state: State, seeds: list[str]) -> bool:
    """Process seeds using Claude agents in parallel with progress tracking."""
    from ..utils import spawn_agents_parallel

    report_dir = Path(state.report_dir)

    # Generate MCP config; select tool preset for URL seeds accordingly
    mcp_config = generate_mcp_config(report_dir)
    if mcp_config:
        url_tools = AGENT_TOOLS["seed_processing"]
    else:
        url_tools = AGENT_TOOLS["seed_processing_fallback"]

    tasks = []

    for i, seed in enumerate(seeds):
        # Skip hidden or unsupported files that slipped through
        if not seed.startswith("http"):
            p = Path(seed)
            if p.name.startswith('.') or (p.suffix.lower() not in _SEED_EXTENSIONS):
                ui.warning(f"Skipping unsupported seed: {p.name}")
                continue

        output_file = report_dir / "full" / "seeds" / f"seed_{i+1}.md"

        # Use brief if available (detailed research instructions), otherwise topic
        research_instructions = state.brief or state.topic

        if seed.startswith("http"):
            if mcp_config:
                fetch_steps = "1. Use mcp__firecrawl__firecrawl_scrape to get the URL content (or mcp__crawl4ai__scrape as backup)"
            else:
                fetch_steps = "1. Use WebFetch to get the URL content"
            prompt = f"""TASK: Fetch URL and save content to file.

URL: {seed}
OUTPUT FILE: {output_file}

STEPS:
{fetch_steps}
2. Extract the main content (title, key text, data)
3. Use Write tool to save to {output_file}

Focus on factual information relevant to: {research_instructions}
Remove navigation, ads, and boilerplate.

CRITICAL: You MUST call the Write tool with file_path="{output_file}" at the end.
"""
        elif seed.lower().endswith(('.xlsx', '.xls')):
            # Excel file - pre-read and pass content directly
            excel_content = _read_excel_file(Path(seed))
            if excel_content:
                prompt = f"""TASK: Summarize Excel data.

OUTPUT FILE: {output_file}

<excel_content>
{excel_content[:50000]}
</excel_content>

Summarize this Excel data, focusing on information relevant to: {research_instructions}

CRITICAL: You MUST call the Write tool with file_path="{output_file}" at the end.
"""
            else:
                ui.warning(f"Could not read Excel file: {seed} (install openpyxl or pandas)")
                continue
        else:
            prompt = f"""TASK: Read file and save summary.

INPUT FILE: {seed}
OUTPUT FILE: {output_file}

STEPS:
1. Use Read tool to read {seed}
2. Summarize key content relevant to: {research_instructions}
3. Use Write tool to save summary to {output_file}

CRITICAL: You MUST call the Write tool with file_path="{output_file}" at the end.
"""

        # URL seeds need fetch tools; file seeds only need Read/Write
        if seed.startswith("http"):
            seed_tools = url_tools
        else:
            seed_tools = ["Read", "Write"]

        tasks.append({
            "id": f"seed_{i+1}",
            "prompt": prompt,
            "model": "sonnet",
            "output_file": str(output_file),
            "timeout_secs": 900,
            "allowed_tools": seed_tools,
            "mcp_config": str(mcp_config) if mcp_config and seed.startswith("http") else None,
        })

    if not tasks:
        return True

    # Progress tracking
    ui.agent_progress_start(len(tasks), "Processing seeds")
    completed = [0]
    lock = threading.Lock()

    def on_complete(task_id: str, result):
        with lock:
            completed[0] += 1
        status = "✓" if result.success else "✗"
        ui.agent_progress_update(completed[0], f"{task_id}: {status}")

    try:
        results = spawn_agents_parallel(tasks, max_workers=5, on_complete=on_complete)
    finally:
        ui.agent_progress_complete(f"Processed {len(tasks)} seeds")
        # Clean up mcp.json (contains API keys)
        if mcp_config:
            try:
                mcp_config.unlink(missing_ok=True)
            except OSError:
                pass

    failed = [tid for tid, r in results.items() if not r.success]
    if failed:
        ui.warning(f"Seed processing failed: {failed}")

    return True


def _summarize_seeds(state: State):
    """Summarize all processed seeds in parallel with progress tracking.

    Reads content directly and passes to summarizer to avoid Read tool failures.
    """
    from ..utils import spawn_agents_parallel

    report_dir = Path(state.report_dir)
    seeds_dir = report_dir / "full" / "seeds"
    summaries_dir = report_dir / "summaries" / "seeds"

    if not seeds_dir.exists():
        return

    tasks = []
    for seed_file in seeds_dir.glob("*.md"):
        summary_file = summaries_dir / seed_file.name

        # Read content here and pass directly
        try:
            content = seed_file.read_text()
        except Exception as e:
            ui.warning(f"Seed file reading failed for {seed_file}: {e}")
            continue

        # Use brief if available (detailed research instructions), otherwise topic
        research_instructions = state.brief or state.topic

        prompt = f"""TASK: Summarize seed reference content.

OUTPUT FILE: {summary_file}

<content>
{content}
</content>

Write a 300-500 word summary including:
- Key findings and data points
- Relevance to topic: {research_instructions}
- Source/citation info if present

CRITICAL: You MUST call Write tool with file_path="{summary_file}" to save your summary.
"""

        tasks.append({
            "id": seed_file.stem,
            "prompt": prompt,
            "model": "sonnet",
            "output_file": str(summary_file),
            "timeout_secs": 360,
            "allowed_tools": ["Write"],
        })

    if not tasks:
        return

    # Progress tracking
    ui.agent_progress_start(len(tasks), "Summarizing seeds")
    completed = [0]
    lock = threading.Lock()

    def on_complete(task_id: str, result):
        with lock:
            completed[0] += 1
        status = "✓" if result.success else "✗"
        ui.agent_progress_update(completed[0], f"{task_id}: {status}")

    results = spawn_agents_parallel(tasks, max_workers=5, on_complete=on_complete)
    ui.agent_progress_complete(f"Summarized {len(tasks)} seeds")

    failed = [tid for tid, r in results.items() if not r.success]
    if failed:
        ui.warning(f"Seed summarization failed: {failed}")


def _write_scope(state: State):
    """Write a scope document based on topic and seeds."""
    report_dir = Path(state.report_dir)
    scope_file = report_dir / "state" / "scope.md"

    # Gather seed summaries if they exist
    seed_context = ""
    summaries_dir = report_dir / "summaries" / "seeds"
    if summaries_dir.exists():
        for f in summaries_dir.glob("*.md"):
            try:
                content = f.read_text()[:2000]  # Limit per seed
                seed_context += f"\n### {f.stem}\n{content}\n"
            except (OSError, IOError) as e:
                ui.warning(f"Seed summary reading failed for {f.name}: {e}")

    # Use brief if available (detailed research instructions), otherwise topic
    research_instructions = state.brief or state.topic

    prompt = f"""TASK: Write research scope document.

Topic: {research_instructions}
Report type: {state.report_type}
Expertise level: {state.expertise_level}
Number of research agents: {state.agent_count}
OUTPUT FILE: {scope_file}

{f"## Seed Material Context{seed_context}" if seed_context else ""}

Include these sections (under 1000 words total):
1. Research objectives (3-5 bullet points)
2. Key questions to answer
3. Boundaries (what's in scope vs out of scope)
4. Expected sections for the final report
5. Quality criteria

CRITICAL: You MUST call Write tool with file_path="{scope_file}" to save the scope document.
"""

    with ui.spinner_task("Writing scope document..."):
        result = spawn_agent(
            prompt, model="opus", output_file=scope_file, timeout_secs=540,
            allowed_tools=["Read", "Write"]
        )
    if not result.success:
        # Write a minimal scope if agent fails
        scope_file.write_text(f"""# Research Scope: {state.topic}

## Objectives
- Provide comprehensive coverage of {state.topic}
- Synthesize current state of knowledge
- Identify key findings and open questions

## Report Type
{state.report_type}

## Target Audience
{state.expertise_level} level readers
""")
